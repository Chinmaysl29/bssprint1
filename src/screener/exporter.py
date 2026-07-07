"""
src/screener/exporter.py
────────────────────────
Excel Exporter for the NIFTY100 Stock Screener.

Public API
──────────
    export_screeners(screener, output_path="output/screener_output.xlsx")
        Run all 6 preset screens, score + sort each result, then write a
        richly-formatted 6-sheet Excel workbook — one sheet per preset.

    ScreenerExporter.export_to_excel(df, filename, ...)
        Low-level single-sheet exporter (backward-compatible).

Output: output/screener_output.xlsx
────────────────────────────────────
    Sheet 1 – Quality Compounder
    Sheet 2 – Value Pick
    Sheet 3 – Growth Accelerator
    Sheet 4 – Dividend Champion
    Sheet 5 – Debt-Free Bluechip
    Sheet 6 – Turnaround Watch

20 KPI columns per sheet (exact order)
────────────────────────────────────────
    1.  company_name      → Company Name
    2.  sector            → Sector
    3.  year              → Year
    4.  roe_percentage    → ROE (%)
    5.  roce_percentage   → ROCE (%)
    6.  npm               → Net Profit Margin (%)
    7.  opm_percentage    → OPM (%)
    8.  debt_equity       → Debt / Equity
    9.  icr               → ICR
    10. fcf               → FCF (₹ Cr)
    11. revenue_cagr_3y   → Revenue CAGR (3Y %)
    12. pat_cagr_3y       → PAT CAGR (3Y %)
    13. eps_cagr_3y       → EPS CAGR (3Y %)
    14. sales             → Sales (₹ Cr)
    15. pat               → Net Profit (₹ Cr)
    16. market_cap        → Market Cap (₹ Cr)
    17. dividend_yield    → Dividend Yield (%)
    18. composite_score   → Composite Score
    19. rank              → Rank
    20. remarks           → Remarks

Score colour bands (Composite Score column)
────────────────────────────────────────────
    ≥ 70  →  green  (S-Tier – Strong Buy)
    50–69 →  yellow (A-Tier – Buy)
    30–49 →  orange (B-Tier – Watch)
    < 30  →  red    (C-Tier – Avoid)

Conditional Formatting (openpyxl CellIsRule)
─────────────────────────────────────────────
    Applied to every numeric KPI column for every data cell:

    KPI              PASS (GREEN)    FAIL (RED)
    ─────────────    ────────────    ──────────
    ROE (%)          ≥ 15            < 15
    ROCE (%)         ≥ 15            < 15
    NPM (%)          ≥ 10            < 10
    OPM (%)          ≥ 15            < 15
    Debt / Equity    ≤ 1.0           > 1.0
    ICR              numeric ≥ 1.5   numeric < 1.5
    FCF (₹ Cr)       > 0             ≤ 0
    Revenue CAGR 3Y  ≥ 10            < 10
    PAT CAGR 3Y      ≥ 10            < 10
    EPS CAGR 3Y      ≥ 10            < 10
    Dividend Yield   ≥ 1.0           < 1.0
    Composite Score  ≥ 50            < 30

    Implementation: cell-by-cell fill applied in Python during write
    (works correctly even for merged/formatted sheets without a live
    Excel engine).  Also registers openpyxl ConditionalFormatting rules
    so the rules persist when the file is re-opened in Excel.
"""

import logging
from pathlib import Path
from typing import TYPE_CHECKING, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from .engine import StockScreener

logger = logging.getLogger(__name__)


# ── Design tokens ─────────────────────────────────────────────────────────────
_BLUE       = "2563EB"
_NAVY       = "0F172A"
_WHITE      = "F8FAFC"
_LIGHT_BLUE = "DBEAFE"   # even row stripe
_LIGHT_GRAY = "F1F5F9"   # odd row stripe
_BORDER_COL = "CBD5E1"

# Score tier fills
_FILL_S = "D1FAE5"   # green-100   ≥ 70
_FILL_A = "FEF9C3"   # yellow-100  50-69
_FILL_B = "FFEDD5"   # orange-100  30-49
_FILL_C = "FEE2E2"   # red-100     < 30

# Remark tier fills (lighter shades used in Remarks column)
_FILL_S_LIGHT = "ECFDF5"
_FILL_A_LIGHT = "FEFCE8"
_FILL_B_LIGHT = "FFF7ED"
_FILL_C_LIGHT = "FEF2F2"

# Conditional-formatting fills (cell-level pass / fail)
_CF_GREEN_PASS = "C6EFCE"   # Excel standard green fill
_CF_RED_FAIL   = "FFC7CE"   # Excel standard red fill
_CF_GREEN_FONT = "276221"   # dark green font
_CF_RED_FONT   = "9C0006"   # dark red font
_CF_YELLOW     = "FFEB9C"   # neutral / warning
_CF_YELLOW_FONT = "9C6500"

# ── Conditional-formatting rule definitions ────────────────────────────────────
#
# Each entry:
#   df_column        – column in the DataFrame
#   pass_operator    – "greaterThanOrEqual" | "lessThanOrEqual" | "greaterThan"
#                      | "lessThan" | "equal" | "between"
#   pass_threshold   – numeric threshold for the PASS (green) rule
#   fail_operator    – operator for the FAIL (red) rule
#   fail_threshold   – numeric threshold for the FAIL (red) rule
#   direction        – "high_good" → higher = greener;
#                      "low_good"  → lower = greener  (D/E, ICR warning)
#
# Rules are applied both as:
#   1. cell-by-cell Python fill (instant, always visible)
#   2. openpyxl ConditionalFormatting registration (preserved in Excel)
#
CF_RULES: list[dict] = [
    # ── Profitability ──────────────────────────────────────────────────────
    {
        "df_col":         "roe_percentage",
        "header":         "ROE (%)",
        "pass_op":        "greaterThanOrEqual",
        "pass_val":       15,
        "fail_op":        "lessThan",
        "fail_val":       15,
        "direction":      "high_good",
        "warn_op":        None,
        "warn_val":       None,
    },
    {
        "df_col":         "roce_percentage",
        "header":         "ROCE (%)",
        "pass_op":        "greaterThanOrEqual",
        "pass_val":       15,
        "fail_op":        "lessThan",
        "fail_val":       15,
        "direction":      "high_good",
        "warn_op":        None,
        "warn_val":       None,
    },
    {
        "df_col":         "npm",
        "header":         "Net Profit Margin (%)",
        "pass_op":        "greaterThanOrEqual",
        "pass_val":       10,
        "fail_op":        "lessThan",
        "fail_val":       10,
        "direction":      "high_good",
        "warn_op":        None,
        "warn_val":       None,
    },
    {
        "df_col":         "opm_percentage",
        "header":         "OPM (%)",
        "pass_op":        "greaterThanOrEqual",
        "pass_val":       15,
        "fail_op":        "lessThan",
        "fail_val":       15,
        "direction":      "high_good",
        "warn_op":        None,
        "warn_val":       None,
    },
    # ── Leverage ───────────────────────────────────────────────────────────
    {
        "df_col":         "debt_equity",
        "header":         "Debt / Equity",
        "pass_op":        "lessThanOrEqual",
        "pass_val":       1.0,
        "fail_op":        "greaterThan",
        "fail_val":       1.0,
        "direction":      "low_good",
        "warn_op":        None,
        "warn_val":       None,
    },
    {
        "df_col":         "icr_numeric",    # derived from icr before inf-replacement
        "header":         "ICR",
        "pass_op":        "greaterThanOrEqual",
        "pass_val":       1.5,
        "fail_op":        "lessThan",
        "fail_val":       1.5,
        "direction":      "high_good",
        "warn_op":        None,
        "warn_val":       None,
    },
    # ── Cash flow ──────────────────────────────────────────────────────────
    {
        "df_col":         "fcf",
        "header":         "FCF (₹ Cr)",
        "pass_op":        "greaterThan",
        "pass_val":       0,
        "fail_op":        "lessThanOrEqual",
        "fail_val":       0,
        "direction":      "high_good",
        "warn_op":        None,
        "warn_val":       None,
    },
    # ── CAGR ───────────────────────────────────────────────────────────────
    {
        "df_col":         "revenue_cagr_3y",
        "header":         "Revenue CAGR (3Y %)",
        "pass_op":        "greaterThanOrEqual",
        "pass_val":       10,
        "fail_op":        "lessThan",
        "fail_val":       10,
        "direction":      "high_good",
        "warn_op":        None,
        "warn_val":       None,
    },
    {
        "df_col":         "pat_cagr_3y",
        "header":         "PAT CAGR (3Y %)",
        "pass_op":        "greaterThanOrEqual",
        "pass_val":       10,
        "fail_op":        "lessThan",
        "fail_val":       10,
        "direction":      "high_good",
        "warn_op":        None,
        "warn_val":       None,
    },
    {
        "df_col":         "eps_cagr_3y",
        "header":         "EPS CAGR (3Y %)",
        "pass_op":        "greaterThanOrEqual",
        "pass_val":       10,
        "fail_op":        "lessThan",
        "fail_val":       10,
        "direction":      "high_good",
        "warn_op":        None,
        "warn_val":       None,
    },
    # ── Yield ──────────────────────────────────────────────────────────────
    {
        "df_col":         "dividend_yield",
        "header":         "Dividend Yield (%)",
        "pass_op":        "greaterThanOrEqual",
        "pass_val":       1.0,
        "fail_op":        "lessThan",
        "fail_val":       1.0,
        "direction":      "high_good",
        "warn_op":        None,
        "warn_val":       None,
    },
    # ── Composite Score ────────────────────────────────────────────────────
    {
        "df_col":         "composite_score",
        "header":         "Composite Score",
        "pass_op":        "greaterThanOrEqual",
        "pass_val":       50,
        "fail_op":        "lessThan",
        "fail_val":       30,
        "direction":      "high_good",
        "warn_op":        None,
        "warn_val":       None,
    },
]

# Build a lookup: df_col → rule dict  (for fast access during cell writing)
_CF_BY_COL: dict[str, dict] = {r["df_col"]: r for r in CF_RULES}

# openpyxl operator strings for CellIsRule
_OP_MAP = {
    "greaterThan":           "greaterThan",
    "greaterThanOrEqual":    "greaterThanOrEqual",
    "lessThan":              "lessThan",
    "lessThanOrEqual":       "lessThanOrEqual",
    "equal":                 "equal",
    "notEqual":              "notEqual",
    "between":               "between",
}


# ── The exact 20 KPI columns ──────────────────────────────────────────────────
# (df_column_name, excel_header, column_width)
KPI_COLUMNS: list[tuple[str, str, int]] = [
    ("company_name",    "Company Name",              30),
    ("sector",          "Sector",                    20),
    ("year",            "Year",                      12),
    ("roe_percentage",  "ROE (%)",                   10),
    ("roce_percentage", "ROCE (%)",                  10),
    ("npm",             "Net Profit Margin (%)",      18),
    ("opm_percentage",  "OPM (%)",                   10),
    ("debt_equity",     "Debt / Equity",              13),
    ("icr",             "ICR",                       10),
    ("fcf",             "FCF (₹ Cr)",                13),
    ("revenue_cagr_3y", "Revenue CAGR (3Y %)",       18),
    ("pat_cagr_3y",     "PAT CAGR (3Y %)",           16),
    ("eps_cagr_3y",     "EPS CAGR (3Y %)",           16),
    ("sales",           "Sales (₹ Cr)",              14),
    ("pat",             "Net Profit (₹ Cr)",          16),
    ("market_cap",      "Market Cap (₹ Cr)",         16),
    ("dividend_yield",  "Dividend Yield (%)",         18),
    ("composite_score", "Composite Score",            16),
    ("rank",            "Rank",                       8),
    ("remarks",         "Remarks",                   22),
]

assert len(KPI_COLUMNS) == 20, "Must be exactly 20 KPI columns"

# Column indices (1-based) for special formatting
_COL_SCORE   = next(i for i, (c, _, _w) in enumerate(KPI_COLUMNS, 1) if c == "composite_score")
_COL_RANK    = next(i for i, (c, _, _w) in enumerate(KPI_COLUMNS, 1) if c == "rank")
_COL_REMARKS = next(i for i, (c, _, _w) in enumerate(KPI_COLUMNS, 1) if c == "remarks")
_COL_COMPANY = next(i for i, (c, _, _w) in enumerate(KPI_COLUMNS, 1) if c == "company_name")


# ── Style helpers ─────────────────────────────────────────────────────────────

def _thin_border() -> Border:
    s = Side(style="thin", color=_BORDER_COL)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_colour: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_colour)


def _score_fill(score: float) -> PatternFill:
    if   score >= 70: return _fill(_FILL_S)
    elif score >= 50: return _fill(_FILL_A)
    elif score >= 30: return _fill(_FILL_B)
    else:             return _fill(_FILL_C)


def _remarks_fill(score: float) -> PatternFill:
    if   score >= 70: return _fill(_FILL_S_LIGHT)
    elif score >= 50: return _fill(_FILL_A_LIGHT)
    elif score >= 30: return _fill(_FILL_B_LIGHT)
    else:             return _fill(_FILL_C_LIGHT)


def _row_fill(row_idx: int) -> PatternFill:
    return _fill(_LIGHT_BLUE if row_idx % 2 == 0 else _LIGHT_GRAY)


def _remarks_text(score: float) -> str:
    if   score >= 70: return "S-Tier – Strong Buy"
    elif score >= 50: return "A-Tier – Buy"
    elif score >= 30: return "B-Tier – Watch"
    else:             return "C-Tier – Avoid"


# ── Conditional-formatting helpers ────────────────────────────────────────────

_CF_PASS_FILL = PatternFill(fill_type="solid", fgColor=_CF_GREEN_PASS)
_CF_FAIL_FILL = PatternFill(fill_type="solid", fgColor=_CF_RED_FAIL)
_CF_WARN_FILL = PatternFill(fill_type="solid", fgColor=_CF_YELLOW)

_CF_PASS_FONT = Font(color=_CF_GREEN_FONT, bold=True, size=10, name="Calibri")
_CF_FAIL_FONT = Font(color=_CF_RED_FONT,   bold=True, size=10, name="Calibri")
_CF_WARN_FONT = Font(color=_CF_YELLOW_FONT, bold=True, size=10, name="Calibri")


def _evaluate_cf(df_col: str, value) -> str:
    """
    Return "pass", "fail", "warn", or "neutral" for a given column value.

    "pass"    → green cell
    "fail"    → red cell
    "warn"    → yellow cell  (currently unused — reserved for future rules)
    "neutral" → no override (fall back to row stripe)
    """
    rule = _CF_BY_COL.get(df_col)
    if rule is None or value is None:
        return "neutral"

    # icr_numeric is stored as float; ICR display column contains string "∞ …"
    try:
        v = float(value)
    except (TypeError, ValueError):
        return "neutral"   # string like "∞ (Debt Free)" → treat as pass for ICR
    if pd.isna(v):
        return "neutral"

    def check(op: str, threshold) -> bool:
        if op == "greaterThanOrEqual":    return v >= threshold
        if op == "greaterThan":           return v >  threshold
        if op == "lessThanOrEqual":       return v <= threshold
        if op == "lessThan":              return v <  threshold
        if op == "equal":                 return v == threshold
        return False

    if check(rule["pass_op"], rule["pass_val"]):
        return "pass"
    if check(rule["fail_op"], rule["fail_val"]):
        return "fail"
    return "neutral"   # between pass and fail → no colour override


def _apply_cf_openpyxl_rules(ws, col_letter: str, first_data_row: int,
                              last_data_row: int, rule: dict) -> None:
    """
    Register openpyxl ConditionalFormatting rules for one column range so
    the rules are preserved and re-evaluated when the file is opened in Excel.
    """
    if first_data_row > last_data_row:
        return

    col_range = f"{col_letter}{first_data_row}:{col_letter}{last_data_row}"

    pass_font_style = Font(color=_CF_GREEN_FONT, bold=True)
    fail_font_style = Font(color=_CF_RED_FONT,   bold=True)

    # PASS rule (green)
    ws.conditional_formatting.add(
        col_range,
        CellIsRule(
            operator=_OP_MAP[rule["pass_op"]],
            formula=[str(rule["pass_val"])],
            fill=PatternFill(fill_type="solid", fgColor=_CF_GREEN_PASS),
            font=pass_font_style,
        ),
    )

    # FAIL rule (red)
    ws.conditional_formatting.add(
        col_range,
        CellIsRule(
            operator=_OP_MAP[rule["fail_op"]],
            formula=[str(rule["fail_val"])],
            fill=PatternFill(fill_type="solid", fgColor=_CF_RED_FAIL),
            font=fail_font_style,
        ),
    )


# ── Data preparation ──────────────────────────────────────────────────────────

def _prepare_df(df: pd.DataFrame, screen_display_name: str) -> pd.DataFrame:
    """
    Guarantee all 20 KPI columns exist, compute derived fields, add Rank
    and Remarks, then return a clean copy sorted by composite_score DESC.
    """
    out = df.copy().reset_index(drop=True)

    # ── Year: use the latest_year from the financial_ratios join if present
    if "year" not in out.columns:
        # fall back to a placeholder — real year comes from the SQL query
        out["year"] = "Latest"

    # ── Net Profit Margin = (net_profit / sales) * 100
    if "npm" not in out.columns:
        if "pat" in out.columns and "sales" in out.columns:
            out["npm"] = out.apply(
                lambda r: round(r["pat"] / r["sales"] * 100, 2)
                if pd.notna(r["pat"]) and pd.notna(r["sales"]) and r["sales"] != 0
                else None,
                axis=1,
            )
        else:
            out["npm"] = None

    # ── ICR: replace inf with a readable string for Excel
    # Keep a numeric shadow column for conditional formatting evaluation
    if "icr" in out.columns:
        out["icr_numeric"] = out["icr"].apply(
            lambda x: 999.0 if x == float("inf") else   # debt-free → treat as very high ICR
                      (float(x) if pd.notna(x) else None)
        )
        out["icr"] = out["icr"].apply(
            lambda x: "∞ (Debt Free)" if x == float("inf") else
                      (round(x, 2) if pd.notna(x) else None)
        )
    else:
        out["icr_numeric"] = None

    # ── Dividend Yield: use market_cap.val5 from the SQL join (already in df)
    if "dividend_yield" not in out.columns:
        # val5 in market_cap table maps to dividend yield
        out["dividend_yield"] = out.get("val5", None)

    # ── Ensure composite_score exists (scored by run_screener)
    if "composite_score" not in out.columns:
        if "advanced_composite_score" in out.columns:
            out["composite_score"] = out["advanced_composite_score"]
        else:
            out["composite_score"] = None

    # ── Sort by composite_score DESC before ranking
    out = out.sort_values("composite_score", ascending=False, na_position="last") \
             .reset_index(drop=True)

    # ── Rank (1-based, by composite_score DESC)
    out["rank"] = range(1, len(out) + 1)

    # ── Remarks (derived from composite_score tier)
    out["remarks"] = out["composite_score"].apply(
        lambda x: _remarks_text(float(x)) if pd.notna(x) else "N/A"
    )

    # ── Round all numeric KPI columns to 2dp
    numeric_cols = [
        "roe_percentage", "roce_percentage", "npm", "opm_percentage",
        "debt_equity", "fcf", "revenue_cagr_3y", "pat_cagr_3y", "eps_cagr_3y",
        "sales", "pat", "market_cap", "dividend_yield", "composite_score",
    ]
    for col in numeric_cols:
        if col in out.columns:
            out[col] = out[col].apply(
                lambda x: round(float(x), 2) if pd.notna(x) else None
            )

    return out


# ── Sheet builder ─────────────────────────────────────────────────────────────

def _write_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame, display_name: str) -> None:
    """
    Write one preset-screen sheet with all 20 KPI columns.

    Conditional formatting is applied at TWO levels:
      1. Cell-by-cell (Python fill) — immediately visible, never overridden.
      2. openpyxl ConditionalFormatting registration — rules re-evaluated
         by Excel when the file is opened (preserves the logic).

    Rules:
        GREEN (pass): value meets the KPI threshold   e.g. ROE >= 15
        RED   (fail): value misses the KPI threshold  e.g. ROE <  15
    """

    ws = wb.create_sheet(title=sheet_name[:31])

    headers = [h for _, h, _ in KPI_COLUMNS]
    n_cols  = len(headers)

    # Build a map  df_col → col_idx (1-based)  for fast CF lookup
    col_idx_map: dict[str, int] = {
        df_col: idx for idx, (df_col, _, _) in enumerate(KPI_COLUMNS, start=1)
    }
    # Also map the ICR display column for CF via icr_numeric
    # (icr display → string; icr_numeric → float used for CF)

    # ── Row 1: Screen title bar ──────────────────────────────────────────────
    title_cell = ws.cell(row=1, column=1, value=f"NIFTY100  ·  {display_name}")
    title_cell.font      = Font(bold=True, size=13, color=_WHITE, name="Calibri")
    title_cell.fill      = _fill(_NAVY)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.row_dimensions[1].height = 28

    # ── Row 2: Sub-title ─────────────────────────────────────────────────────
    count_text = (
        f"{len(df)} companies  |  Sorted by Composite Score DESC  |  "
        "GREEN = Condition Passed  ·  RED = Condition Failed"
    )
    sub_cell   = ws.cell(row=2, column=1, value=count_text)
    sub_cell.font      = Font(italic=True, size=9, color="64748B", name="Calibri")
    sub_cell.fill      = _fill("F8FAFC")
    sub_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.row_dimensions[2].height = 16

    # ── Row 3: Column headers ────────────────────────────────────────────────
    for col_idx, header in enumerate(headers, start=1):
        cell           = ws.cell(row=3, column=col_idx, value=header)
        cell.font      = Font(bold=True, color=_WHITE, size=10, name="Calibri")
        cell.fill      = _fill(_BLUE)
        cell.border    = _thin_border()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 36

    FIRST_DATA_ROW = 4
    last_data_row  = FIRST_DATA_ROW + len(df) - 1

    # ── Rows 4+: Data + cell-by-cell conditional formatting ──────────────────
    for row_offset, (_, data_row) in enumerate(df.iterrows()):
        excel_row = FIRST_DATA_ROW + row_offset
        ws.row_dimensions[excel_row].height = 18

        # Composite score → used for Score/Remarks columns
        score_val = data_row.get("composite_score", None)
        try:
            score_f = float(score_val) if pd.notna(score_val) else None
        except (TypeError, ValueError):
            score_f = None

        base_fill = _row_fill(row_offset + 1)

        for col_idx, (df_col, _header, _width) in enumerate(KPI_COLUMNS, start=1):

            # ── Fetch display value ──────────────────────────────────────
            if df_col in df.columns:
                raw   = data_row[df_col]
                value = None if pd.isna(raw) else (
                    str(raw).strip() if isinstance(raw, str) else raw
                )
            else:
                value = None

            cell        = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.border = _thin_border()

            # ── Determine the CF evaluation value ────────────────────────
            # For the ICR display column we use icr_numeric for CF logic
            if df_col == "icr":
                cf_eval_col = "icr_numeric"
                cf_eval_val = data_row.get("icr_numeric", None)
            else:
                cf_eval_col = df_col
                cf_eval_val = value

            cf_result = _evaluate_cf(cf_eval_col, cf_eval_val)

            # ── Apply fills & fonts ───────────────────────────────────────
            if col_idx == _COL_SCORE and score_f is not None:
                # Composite Score: tier-based 4-colour fill (overrides CF)
                cell.fill  = _score_fill(score_f)
                cell.font  = Font(bold=True, size=10, name="Calibri", color=_NAVY)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            elif col_idx == _COL_RANK:
                cell.fill  = base_fill
                cell.font  = Font(bold=True, size=10, name="Calibri", color="475569")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            elif col_idx == _COL_REMARKS and score_f is not None:
                cell.fill  = _remarks_fill(score_f)
                cell.font  = Font(italic=True, size=9, name="Calibri", color=_NAVY)
                cell.alignment = Alignment(horizontal="left", vertical="center")

            elif col_idx == _COL_COMPANY:
                cell.fill  = base_fill
                cell.font  = Font(bold=True, size=10, name="Calibri")
                cell.alignment = Alignment(horizontal="left", vertical="center")

            elif cf_result == "pass":
                # ── GREEN: condition passed ──────────────────────────────
                cell.fill  = _CF_PASS_FILL
                cell.font  = _CF_PASS_FONT
                cell.alignment = Alignment(horizontal="right", vertical="center")

            elif cf_result == "fail":
                # ── RED: condition failed ────────────────────────────────
                cell.fill  = _CF_FAIL_FILL
                cell.font  = _CF_FAIL_FONT
                cell.alignment = Alignment(horizontal="right", vertical="center")

            else:
                # Neutral / text columns
                cell.fill  = base_fill
                cell.font  = Font(size=10, name="Calibri")
                is_num = isinstance(value, (int, float))
                cell.alignment = Alignment(
                    horizontal="right" if is_num else "left",
                    vertical="center",
                )

    # ── Register openpyxl ConditionalFormatting rules (Excel-native) ─────────
    # Applied per numeric KPI column so rules survive re-open in Excel.
    # Note: ICR display column contains strings ("∞ (Debt Free)", numbers);
    #       we skip registering CF for the ICR display column because Excel
    #       can't evaluate string cells — the Python cell-fill handles it.
    if last_data_row >= FIRST_DATA_ROW:
        for rule in CF_RULES:
            df_col = rule["df_col"]
            if df_col == "icr_numeric":
                # Map back to the ICR display column index for Excel rules
                # (icr_numeric is not a KPI column — skip Excel-side CF for ICR)
                continue
            if df_col not in col_idx_map:
                continue
            col_letter = get_column_letter(col_idx_map[df_col])
            _apply_cf_openpyxl_rules(
                ws, col_letter, FIRST_DATA_ROW, last_data_row, rule
            )

    # ── Column widths ─────────────────────────────────────────────────────────
    for col_idx, (_df_col, _header, width) in enumerate(KPI_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Freeze panes at row 4 ────────────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=4, column=1)

    # ── Autofilter on header row ──────────────────────────────────────────────
    if last_data_row >= FIRST_DATA_ROW:
        ws.auto_filter.ref = (
            f"A3:{get_column_letter(n_cols)}{last_data_row}"
        )


# ── Public entry point ────────────────────────────────────────────────────────

def export_screeners(
    screener: "StockScreener",
    output_path: str = "output/screener_output.xlsx",
) -> Path:
    """
    Run all 6 preset screens and write ``output/screener_output.xlsx`` with
    one sheet per preset, each containing exactly 20 KPI columns.

    Pipeline per sheet
    ──────────────────
        Load Data  →  Apply Preset Filter  →  Calculate Score
          →  Sort Score DESC  →  Add Rank & Remarks  →  Write Sheet

    Parameters
    ──────────
    screener    : Initialised StockScreener instance
    output_path : Destination file (default: ``output/screener_output.xlsx``)

    Returns
    ───────
    Path : Resolved path to the written workbook.

    Example
    ───────
    >>> from src.screener.engine import StockScreener
    >>> from src.screener.exporter import export_screeners
    >>> path = export_screeners(StockScreener())
    >>> print(path)
    .../output/screener_output.xlsx
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # ── 1. Collect preset names ───────────────────────────────────────────────
    preset_names: list[str] = list(screener.config.get("screens", {}).keys())
    if not preset_names:
        raise ValueError(
            "No screens in screener_config.yaml. "
            "Add at least one entry under 'screens'."
        )
    logger.info("export_screeners: %d presets → %s", len(preset_names), output_path)

    # ── 2. Need dividend_yield from market_cap.val5 ───────────────────────────
    # Extend the SQL in engine to include val5 if not already present.
    # We do this by patching the cached financial_data after the fact.
    _ensure_dividend_yield(screener)

    # ── 3. Run all presets & write workbook ───────────────────────────────────
    wb = Workbook()
    # Remove default empty sheet
    default = wb.active
    if default is not None:
        wb.remove(default)

    for preset_key in preset_names:
        result       = screener.run_screener(screen_name=preset_key)
        display_name = result["screen_display_name"]
        companies_df = pd.DataFrame(result["companies"])

        # Add dividend_yield to this slice if available
        if "dividend_yield" not in companies_df.columns and screener.financial_data is not None:
            dy_map = screener.financial_data.set_index("company_id")["dividend_yield"] \
                     if "dividend_yield" in screener.financial_data.columns else None
            if dy_map is not None:
                companies_df["dividend_yield"] = companies_df["company_id"].map(dy_map)

        # Derive year from the loaded financial_data
        if "year" not in companies_df.columns and screener.financial_data is not None:
            if "year" in screener.financial_data.columns:
                yr_map = screener.financial_data.set_index("company_id")["year"]
                companies_df["year"] = companies_df["company_id"].map(yr_map)

        prepared = _prepare_df(companies_df, display_name)
        _write_sheet(wb, display_name, prepared, display_name)
        logger.debug("  wrote sheet '%s' (%d rows)", display_name, len(prepared))

    wb.save(output_path)
    logger.info(
        "export_screeners: saved %s  (%d sheets)",
        output_path, len(wb.sheetnames)
    )
    return output_path.resolve()


def _ensure_dividend_yield(screener: "StockScreener") -> None:
    """
    Load dividend_yield (market_cap.val5) into screener.financial_data
    if it is not already present.
    """
    if screener.financial_data is None:
        screener.load_financial_data()

    if "dividend_yield" in screener.financial_data.columns:
        return  # already there

    try:
        import sqlite3, pandas as pd
        conn = sqlite3.connect(screener.db_path)
        dy = pd.read_sql(
            """
            SELECT mc.company_id, mc.val5 as dividend_yield
            FROM market_cap mc
            JOIN (
                SELECT company_id, MAX(year) as max_yr
                FROM market_cap GROUP BY company_id
            ) mx ON mc.company_id = mx.company_id AND mc.year = mx.max_yr
            """,
            conn,
        )
        conn.close()
        screener.financial_data = screener.financial_data.merge(
            dy, on="company_id", how="left"
        )
        logger.debug("_ensure_dividend_yield: merged val5 → dividend_yield")
    except Exception as exc:
        logger.warning("_ensure_dividend_yield failed: %s", exc)
        screener.financial_data["dividend_yield"] = None


# ── ScreenerExporter class (backward-compatible) ──────────────────────────────

class ScreenerExporter:
    """
    Low-level single-DataFrame exporter, retained for backward-compatibility
    with ``StockScreener.run_screener(export_to_excel=True)`` and existing tests.

    For the full 6-sheet export use ``export_screeners()`` instead.
    """

    @staticmethod
    def export_to_excel(
        df: pd.DataFrame,
        filename: str = "screener_output.xlsx",
        sort_by: Optional[str] = "advanced_composite_score",
        ascending: bool = False,
    ) -> bool:
        """Export a single scored DataFrame to Excel (one sheet)."""
        if df is None or df.empty:
            return False

        sorted_df = df.copy()
        if sort_by and sort_by in sorted_df.columns:
            sorted_df = sorted_df.sort_values(
                by=sort_by, ascending=ascending, na_position="last"
            )

        primary_cols = [
            "company_id", "company_name", "sector", "market_cap",
            "roe_percentage", "roce_percentage", "debt_equity",
            "revenue_cagr_3y", "revenue_cagr_5y",
            "pat_cagr_3y", "pat_cagr_5y",
            "cfo_quality_score", "opm_percentage", "asset_turnover",
            "advanced_composite_score", "composite_quality_score",
        ]
        sector_cols = [c for c in sorted_df.columns if c.endswith("_vs_sector")]
        export_cols = [c for c in (primary_cols + sector_cols) if c in sorted_df.columns]
        export_df   = sorted_df[export_cols]

        output_path = Path(filename)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            export_df.to_excel(writer, sheet_name="Screener Results", index=False)
            ws = writer.sheets["Screener Results"]

            hdr_fill = PatternFill(fill_type="solid", fgColor=_BLUE)
            border   = _thin_border()
            for cell in ws[1]:
                cell.font      = Font(bold=True, color=_WHITE, size=11, name="Calibri")
                cell.fill      = hdr_fill
                cell.border    = border
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for col_idx, col in enumerate(export_df.columns, start=1):
                col_letter = get_column_letter(col_idx)
                try:
                    max_len = max(
                        export_df[col].astype(str).map(len).max()
                        if not export_df[col].isna().all() else len(col),
                        len(col),
                    ) + 2
                except Exception:
                    max_len = len(col) + 2
                ws.column_dimensions[col_letter].width = min(max_len, 50)

            ws.freeze_panes = "A2"

        return True


# ── Column utility (kept for any legacy callers) ──────────────────────────────

def column_index_to_letter(index: int) -> str:
    """0-based column index → Excel column letter."""
    return get_column_letter(index + 1)
