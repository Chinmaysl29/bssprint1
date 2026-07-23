"""Generate peer comparison Excel reports from peer percentile data."""

from __future__ import annotations

import re
import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.analytics.peer import get_database_path


BASE_DIR = Path(__file__).resolve().parents[2]
DEFAULT_OUTPUT_PATH = BASE_DIR / "output" / "peer_comparison.xlsx"

REPORT_COLUMNS = [
    "company_id",
    "company_name",
    "sector",
    "year",
    "ROE",
    "ROE_percentile",
    "ROCE",
    "ROCE_percentile",
    "Net Profit Margin",
    "Net_Profit_Margin_percentile",
    "Operating Margin",
    "Debt Equity",
    "Debt_percentile",
    "ICR",
    "ICR_percentile",
    "FCF",
    "FCF_percentile",
    "CFO Quality",
    "Revenue CAGR",
    "PAT CAGR",
    "EPS CAGR",
    "EPS_percentile",
    "Asset Turnover",
    "Asset_Turnover_percentile",
    "Sales",
    "Net Profit",
    "Market Cap",
    "Dividend Yield",
    "Dividend_Payout_percentile",
    "Composite Score",
    "Quality Rank",
    "Risk Score",
    "Total_Debt_percentile",
    "Peer Rank",
]

NUMERIC_COLUMNS = {
    "year",
    "ROE",
    "ROE_percentile",
    "ROCE",
    "ROCE_percentile",
    "Net Profit Margin",
    "Net_Profit_Margin_percentile",
    "Operating Margin",
    "Debt Equity",
    "Debt_percentile",
    "ICR",
    "ICR_percentile",
    "FCF",
    "FCF_percentile",
    "CFO Quality",
    "Revenue CAGR",
    "PAT CAGR",
    "EPS CAGR",
    "EPS_percentile",
    "Asset Turnover",
    "Asset_Turnover_percentile",
    "Sales",
    "Net Profit",
    "Market Cap",
    "Dividend Yield",
    "Dividend_Payout_percentile",
    "Composite Score",
    "Quality Rank",
    "Risk Score",
    "Total_Debt_percentile",
    "Peer Rank",
}

PERCENTILE_METRICS = {
    "ROE": "ROE_percentile",
    "ROCE": "ROCE_percentile",
    "Net_Profit_Margin": "Net_Profit_Margin_percentile",
    "D/E": "Debt_percentile",
    "Interest_Coverage": "ICR_percentile",
    "FCF": "FCF_percentile",
    "EPS": "EPS_percentile",
    "Asset_Turnover": "Asset_Turnover_percentile",
    "Dividend_Payout": "Dividend_Payout_percentile",
    "Total_Debt": "Total_Debt_percentile",
}

REPORT_SHEETS = [
    "IT Services",
    "Banking",
    "FMCG",
    "Automobile",
    "Pharma",
    "Energy",
    "Metals",
    "Telecom",
    "Consumer",
    "Finance",
    "Others",
]

PEER_GROUP_TO_SHEET = {
    "IT Services": "IT Services",
    "Private Banks": "Banking",
    "Public Sector Banks": "Banking",
    "FMCG": "FMCG",
    "Automobiles": "Automobile",
    "Pharmaceuticals": "Pharma",
    "Oil & Gas": "Energy",
    "Power & Utilities": "Energy",
    "Steel": "Metals",
    "Consumer Finance": "Finance",
    "Life Insurance": "Finance",
}

SECTOR_TO_SHEET_FALLBACK = {
    "Communication Services": "Telecom",
    "Consumer Discretionary": "Consumer",
    "Financials": "Finance",
    "Energy": "Energy",
    "Healthcare": "Pharma",
    "Materials": "Metals",
}


def load_peer_data(db_path: str | Path | None = None) -> pd.DataFrame:
    """Load company-level KPI data and map each company to a report sheet."""
    database_path = Path(db_path) if db_path is not None else Path(get_database_path())

    with sqlite3.connect(database_path) as conn:
        ratios_df = pd.read_sql_query(
            """
            SELECT
                company_id,
                year AS ratio_year,
                ratio1 AS net_profit_margin,
                ratio2 AS operating_margin,
                ratio3 AS ratio_roe,
                ratio4 AS debt_equity_ratio,
                ratio5 AS icr_ratio,
                ratio6 AS asset_turnover_ratio,
                free_cash_flow_cr AS fcf_ratio,
                cfo_quality_score AS cfo_quality_ratio
            FROM financial_ratios
            """,
            conn,
        )
        percentiles_df = pd.read_sql_query(
            """
            SELECT company_id, peer_group_name, year, metric, percentile_rank
            FROM peer_percentiles
            """,
            conn,
        )

    from src.screener.engine import StockScreener

    screener = StockScreener(
        db_path=str(database_path),
        config_path=str(BASE_DIR / "config" / "screener_config.yaml"),
    )
    try:
        financial_df = screener.load_financial_data(compute_scores=True)
    finally:
        screener.close()

    if financial_df.empty:
        return pd.DataFrame(columns=REPORT_COLUMNS)

    ratios_df["ratio_year_int"] = ratios_df["ratio_year"].apply(_year_to_int)
    ratios_df = ratios_df.sort_values(["company_id", "ratio_year_int"]).drop_duplicates(
        "company_id",
        keep="last",
    )
    peer_map = _latest_peer_groups(percentiles_df)
    percentile_wide = _latest_peer_percentiles(percentiles_df)

    df = financial_df.merge(ratios_df, on="company_id", how="left")
    df = df.merge(peer_map, on="company_id", how="left")
    df = df.merge(percentile_wide, on="company_id", how="left")
    df["report_sheet"] = _assign_report_sheet(df)

    report_df = pd.DataFrame()
    report_df["company_id"] = df["company_id"]
    report_df["company_name"] = df["company_name"]
    report_df["sector"] = df["sector"]
    report_df["year"] = df["year"]
    report_df["ROE"] = _coalesce(df, "ratio_roe", "roe_percentage")
    report_df["ROE_percentile"] = df["ROE_percentile"]
    report_df["ROCE"] = df["roce_percentage"]
    report_df["ROCE_percentile"] = df["ROCE_percentile"]
    report_df["Net Profit Margin"] = df["net_profit_margin"]
    report_df["Net_Profit_Margin_percentile"] = df["Net_Profit_Margin_percentile"]
    report_df["Operating Margin"] = _coalesce(df, "operating_margin", "opm_percentage")
    report_df["Debt Equity"] = _coalesce(df, "debt_equity_ratio", "debt_equity")
    report_df["Debt_percentile"] = df["Debt_percentile"]
    report_df["ICR"] = _coalesce(df, "icr", "icr_ratio").replace([float("inf"), -float("inf")], pd.NA)
    report_df["ICR_percentile"] = df["ICR_percentile"]
    report_df["FCF"] = _coalesce(df, "fcf_ratio", "fcf")
    report_df["FCF_percentile"] = df["FCF_percentile"]
    report_df["CFO Quality"] = _coalesce(df, "cfo_quality_ratio", "cfo_quality_score", "avg_cfo_pat_ratio")
    report_df["Revenue CAGR"] = _coalesce(df, "revenue_cagr_5y", "revenue_cagr_3y")
    report_df["PAT CAGR"] = _coalesce(df, "pat_cagr_5y", "pat_cagr_3y")
    report_df["EPS CAGR"] = _coalesce(df, "eps_cagr_5y", "eps_cagr_3y")
    report_df["EPS_percentile"] = df["EPS_percentile"]
    report_df["Asset Turnover"] = _coalesce(df, "asset_turnover_ratio", "asset_turnover")
    report_df["Asset_Turnover_percentile"] = df["Asset_Turnover_percentile"]
    report_df["Sales"] = df["sales"]
    report_df["Net Profit"] = df["pat"]
    report_df["Market Cap"] = df["market_cap"]
    report_df["Dividend Yield"] = df["dividend_yield"]
    report_df["Dividend_Payout_percentile"] = df["Dividend_Payout_percentile"]
    report_df["Composite Score"] = df["composite_score"]
    report_df["__benchmark_score"] = df["composite_quality_score"]
    report_df["Quality Rank"] = df["sector_quality_rank"]
    report_df["Risk Score"] = 100 - (pd.to_numeric(df["score_leverage"], errors="coerce") / 15 * 100)
    report_df["Total_Debt_percentile"] = df["Total_Debt_percentile"]
    report_df["report_sheet"] = df["report_sheet"]

    report_df["Peer Rank"] = report_df.groupby("report_sheet")["__benchmark_score"].rank(
        ascending=False,
        method="min",
        na_option="bottom",
    )
    report_df = _fill_missing_percentiles(report_df)
    report_df = report_df.reindex(columns=REPORT_COLUMNS + ["report_sheet"])

    for column in NUMERIC_COLUMNS:
        report_df[column] = pd.to_numeric(report_df[column], errors="coerce")

    report_df["year"] = report_df["year"].astype("Int64")
    report_df["Quality Rank"] = report_df["Quality Rank"].astype("Int64")
    report_df["Peer Rank"] = report_df["Peer Rank"].astype("Int64")

    return report_df.sort_values(
        ["report_sheet", "Peer Rank", "company_name"],
        na_position="last",
    ).reset_index(drop=True)


def add_peer_median(peer_df: pd.DataFrame) -> pd.DataFrame:
    """Return a copy of peer data.

    The public function is retained for the report generator API, but the
    workbook schema is fixed by REPORT_COLUMNS and should not add extra fields.
    """
    return peer_df.copy()


def generate_peer_sheets(peer_df: pd.DataFrame | None = None) -> dict[str, pd.DataFrame]:
    """Return the exact 11 report sheets keyed by Excel sheet name."""
    df = add_peer_median(peer_df if peer_df is not None else load_peer_data())
    if "report_sheet" not in df.columns:
        df["report_sheet"] = "Others"

    sheets: dict[str, pd.DataFrame] = {}

    for sheet_name in REPORT_SHEETS:
        group_df = df[df["report_sheet"].eq(sheet_name)]
        sheet_df = group_df.reindex(columns=REPORT_COLUMNS).sort_values(
            ["Peer Rank", "company_name"],
            na_position="last",
        )
        sheet_df = _append_peer_median_row(sheet_df)
        sheets[sheet_name] = sheet_df.reset_index(drop=True)

    return sheets


def apply_formatting(workbook) -> None:
    """Apply consistent formatting to all peer report worksheets."""
    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(color="F8FAFC", bold=True)
    stripe_fill = PatternFill("solid", fgColor="F8FAFC")
    median_fill = PatternFill("solid", fgColor="DBEAFE")
    benchmark_fill = PatternFill("solid", fgColor="FFD966")
    peer_median_fill = PatternFill("solid", fgColor="FCD34D")
    percentile_green = PatternFill("solid", fgColor="C6EFCE")
    percentile_yellow = PatternFill("solid", fgColor="FFEB9C")
    percentile_red = PatternFill("solid", fgColor="FFC7CE")
    border = Border(bottom=Side(style="thin", color="CBD5E1"))

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        headers = {cell.value: idx for idx, cell in enumerate(worksheet[1], start=1)}
        percentile_col = headers.get("percentile_rank")
        peer_rank_col = headers.get("Peer Rank")
        company_name_col = headers.get("company_name")
        median_cols = {
            headers[column]
            for column in headers
            if isinstance(column, str) and column.endswith("Peer Median")
        }
        rank_cols = set()
        percentile_cols = {
            headers[column]
            for column in headers
            if isinstance(column, str) and column.endswith("_percentile")
        }

        for row in worksheet.iter_rows(min_row=2):
            if row[0].row % 2 == 0:
                for cell in row:
                    cell.fill = stripe_fill

            is_peer_median_row = (
                company_name_col is not None
                and row[company_name_col - 1].value == "Peer Median"
            )

            if is_peer_median_row:
                for cell in row:
                    cell.fill = peer_median_fill
                    cell.font = Font(bold=True, color="0F172A")

            if peer_rank_col is not None and not is_peer_median_row:
                try:
                    peer_rank = float(row[peer_rank_col - 1].value)
                except (TypeError, ValueError):
                    peer_rank = None
                if peer_rank == 1:
                    for cell in row:
                        cell.fill = benchmark_fill

            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top")
                header = worksheet.cell(row=1, column=cell.column).value
                if header in NUMERIC_COLUMNS:
                    cell.number_format = "0.00"

            for col_idx in median_cols:
                row[col_idx - 1].fill = median_fill

            for col_idx in rank_cols or ([percentile_col] if percentile_col is not None else []):
                cell = row[col_idx - 1]
                try:
                    percentile = float(cell.value)
                except (TypeError, ValueError):
                    percentile = None
                if percentile is not None:
                    if percentile >= 75:
                        cell.fill = PatternFill("solid", fgColor="D1FAE5")
                    elif percentile < 25:
                        cell.fill = PatternFill("solid", fgColor="FEE2E2")

            for col_idx in percentile_cols:
                cell = row[col_idx - 1]
                if is_peer_median_row:
                    continue
                try:
                    percentile = float(cell.value)
                except (TypeError, ValueError):
                    percentile = None
                if percentile is None:
                    continue
                if percentile >= 75:
                    cell.fill = percentile_green
                elif percentile <= 25:
                    cell.fill = percentile_red
                else:
                    cell.fill = percentile_yellow

        for col_idx, column_cells in enumerate(worksheet.columns, start=1):
            values = [str(cell.value) for cell in column_cells if cell.value is not None]
            width = min(max([len(value) for value in values] + [10]) + 2, 34)
            worksheet.column_dimensions[get_column_letter(col_idx)].width = width


def export_excel(
    output_path: str | Path = DEFAULT_OUTPUT_PATH,
    db_path: str | Path | None = None,
) -> Path:
    """Generate output/peer_comparison.xlsx and return its path."""
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    peer_df = add_peer_median(load_peer_data(db_path=db_path))
    sheets = generate_peer_sheets(peer_df)

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        if not sheets:
            pd.DataFrame(columns=REPORT_COLUMNS).to_excel(
                writer,
                sheet_name="Peer Comparison",
                index=False,
            )
        else:
            for sheet_name, sheet_df in sheets.items():
                sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)

        apply_formatting(writer.book)

    return output_file


def _safe_sheet_name(name: str, used_names: set[str]) -> str:
    """Return an Excel-safe, unique sheet name."""
    clean_name = re.sub(r"[\[\]:*?/\\]", " ", name).strip() or "Sheet"
    clean_name = re.sub(r"\s+", " ", clean_name)[:31]

    candidate = clean_name
    suffix = 1
    while candidate in used_names:
        suffix_text = f" {suffix}"
        candidate = f"{clean_name[:31 - len(suffix_text)]}{suffix_text}"
        suffix += 1

    return candidate


def _latest_peer_groups(percentiles_df: pd.DataFrame) -> pd.DataFrame:
    """Return each company's latest peer group from peer_percentiles."""
    if percentiles_df.empty:
        return pd.DataFrame(columns=["company_id", "peer_group_name"])

    df = percentiles_df.copy()
    df["year"] = pd.to_numeric(df["year"], errors="coerce")
    df = df.sort_values(["company_id", "year"], na_position="first")
    return df.drop_duplicates("company_id", keep="last")[["company_id", "peer_group_name"]]


def _assign_report_sheet(df: pd.DataFrame) -> pd.Series:
    """Assign target report sheet from peer group, falling back to broad sector."""
    report_sheet = df["peer_group_name"].map(PEER_GROUP_TO_SHEET)
    fallback_sheet = df["sector"].map(SECTOR_TO_SHEET_FALLBACK)
    return report_sheet.combine_first(fallback_sheet).fillna("Others")


def _fill_missing_percentiles(report_df: pd.DataFrame) -> pd.DataFrame:
    """Fill missing percentile columns using within-sheet ranks."""
    df = report_df.copy()
    metric_pairs = {
        "ROE_percentile": ("ROE", True),
        "ROCE_percentile": ("ROCE", True),
        "Net_Profit_Margin_percentile": ("Net Profit Margin", True),
        "Debt_percentile": ("Debt Equity", False),
        "ICR_percentile": ("ICR", True),
        "FCF_percentile": ("FCF", True),
        "EPS_percentile": ("EPS CAGR", True),
        "Asset_Turnover_percentile": ("Asset Turnover", True),
        "Dividend_Payout_percentile": ("Dividend Yield", True),
        "Total_Debt_percentile": ("Debt Equity", False),
    }

    for percentile_col, (metric_col, high_is_good) in metric_pairs.items():
        if percentile_col not in df.columns or metric_col not in df.columns:
            continue
        computed = df.groupby("report_sheet")[metric_col].transform(
            lambda values: pd.to_numeric(values, errors="coerce").rank(
                pct=True,
                method="average",
                ascending=high_is_good,
            )
            * 100
        )
        if high_is_good:
            fallback = computed
        else:
            fallback = 100 - computed
        df[percentile_col] = pd.to_numeric(df[percentile_col], errors="coerce").combine_first(
            fallback.round(2)
        )

    return df


def _append_peer_median_row(sheet_df: pd.DataFrame) -> pd.DataFrame:
    """Append one Peer Median row to the bottom of a sheet dataframe."""
    median_row = {column: pd.NA for column in REPORT_COLUMNS}
    median_row["company_name"] = "Peer Median"

    if not sheet_df.empty:
        for column in NUMERIC_COLUMNS:
            if column in {"year", "Quality Rank", "Peer Rank"}:
                continue
            if column in sheet_df.columns:
                values = pd.to_numeric(sheet_df[column], errors="coerce").dropna()
                if not values.empty:
                    median_row[column] = values.median()

    result_df = sheet_df.reset_index(drop=True).copy()
    result_df.loc[len(result_df), REPORT_COLUMNS] = [median_row[column] for column in REPORT_COLUMNS]
    return result_df


def _latest_peer_percentiles(percentiles_df: pd.DataFrame) -> pd.DataFrame:
    """Return latest requested percentile metrics as one row per company."""
    if percentiles_df.empty:
        return pd.DataFrame(columns=["company_id", *PERCENTILE_METRICS.values()])

    df = percentiles_df[percentiles_df["metric"].isin(PERCENTILE_METRICS)].copy()
    if df.empty:
        return pd.DataFrame(columns=["company_id", *PERCENTILE_METRICS.values()])

    df["year"] = pd.to_numeric(df["year"], errors="coerce")
    latest_year = df.groupby("company_id")["year"].transform("max")
    df = df[df["year"].eq(latest_year)].copy()

    wide_df = df.pivot_table(
        index="company_id",
        columns="metric",
        values="percentile_rank",
        aggfunc="first",
    ).rename(columns=PERCENTILE_METRICS)

    return wide_df.reindex(columns=list(PERCENTILE_METRICS.values())).reset_index()


def _coalesce(df: pd.DataFrame, *columns: str) -> pd.Series:
    """Return the first non-null value across available columns."""
    result = None
    for column in columns:
        if column in df.columns:
            if result is None:
                result = df[column].copy()
            else:
                result = result.combine_first(df[column])

    if result is None:
        return pd.Series(pd.NA, index=df.index)
    return result


def _year_to_int(year_value) -> int | None:
    """Extract a 4-digit year from values like 'Mar 2024'."""
    if pd.isna(year_value):
        return None

    match = re.search(r"(19|20)\d{2}", str(year_value))
    return int(match.group(0)) if match else None


if __name__ == "__main__":
    print(export_excel())
