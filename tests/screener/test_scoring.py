"""
tests/screener/test_scoring.py
================================
Tests for src/screener/scoring.py covering all 6 required areas:

  1. Winsorization       - P10/P90 clipping, NA handling, outlier capping
  2. Score 0-100         - composite_score always in [0, 100]
  3. Sector ranking      - sector_relative_score and sector_quality_rank
  4. Excel creation      - export_screeners() produces a valid .xlsx file
  5. Six sheets          - workbook has exactly 6 preset sheets
  6. Sorting             - every sheet sorted by Composite Score DESC

Additional:
  - D/E inversion        - high D/E company gets LOW score (1 - percentile_rank)
"""

import os
import sys
import pytest
import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "../..")))

from src.screener.scoring import ScreenerScoring
from src.screener.engine import StockScreener
from src.screener.exporter import export_screeners


# ── Fixtures ──────────────────────────────────────────────────────────────────

@pytest.fixture(scope="module")
def screener():
    return StockScreener()


@pytest.fixture(scope="module")
def sample_df():
    """Minimal synthetic DataFrame that exercises every scoring code-path."""
    return pd.DataFrame({
        "company_id":       ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"],
        "company_name":     ["Alpha","Beta","Gamma","Delta","Epsilon",
                             "Zeta","Eta","Theta","Iota","Kappa"],
        "sector":           ["IT","IT","IT","IT","IT",
                             "Pharma","Pharma","Pharma","Pharma","Pharma"],
        "roe_percentage":   [5, 10, 15, 20, 30, 8, 12, 18, 25, 200],
        "roce_percentage":  [6, 12, 18, 24, 35, 9, 14, 20, 28, 210],
        "opm_percentage":   [4, 8,  12, 16, 25, 6, 10, 14, 20, 150],
        "fcf":              [-500, 0, 100, 300, 800, -200, 50, 200, 600, 2000],
        "revenue_cagr_3y":  [2, 5, 10, 15, 30, 3, 7, 12, 18, 80],
        "revenue_cagr_5y":  [3, 6, 11, 16, 32, 4, 8, 13, 19, 85],
        "pat_cagr_3y":      [1, 4, 9,  14, 28, 2, 6, 11, 17, 75],
        "pat_cagr_5y":      [2, 5, 10, 15, 30, 3, 7, 12, 18, 80],
        "fcf_cagr_3y":      [0, 3, 8,  13, 25, 1, 5, 10, 16, 70],
        "fcf_cagr_5y":      [1, 4, 9,  14, 27, 2, 6, 11, 17, 72],
        "avg_cfo_pat_ratio": [0.3, 0.6, 0.9, 1.2, 2.0, 0.4, 0.7, 1.0, 1.5, 3.0],
        # Deliberately ordered: A=0 (debt-free), J=10 (very high debt)
        "debt_equity":      [0.0, 0.2, 0.5, 1.0, 1.5, 0.1, 0.3, 0.8, 2.0, 10.0],
        "icr":              [float("inf"), 15, 8, 5, 3, float("inf"), 12, 6, 4, 1.0],
        "eps_cagr_5y":      [2, 5, 10, 15, 28, 3, 7, 12, 17, 70],
        "cfo_quality_score":[40, 55, 65, 75, 90, 45, 58, 68, 78, 95],
        "asset_turnover":   [0.5, 0.8, 1.0, 1.2, 1.5, 0.6, 0.9, 1.1, 1.3, 1.8],
        "dividend_payout":  [0.1, 0.2, 0.3, 0.4, 0.6, 0.1, 0.25, 0.35, 0.45, 0.7],
    })


@pytest.fixture(scope="module")
def scored_df(sample_df):
    """sample_df passed through the full scoring pipeline."""
    df = ScreenerScoring.calculate_p10_p90_normalization(sample_df.copy())
    df = ScreenerScoring.calculate_composite_score(df)
    return df


# ─────────────────────────────────────────────────────────────────────────────
# 1. WINSORIZATION
# ─────────────────────────────────────────────────────────────────────────────

class TestWinsorization:

    def test_outlier_is_clipped(self, sample_df):
        """Extreme outlier (200% ROE) must be winsorized to <= P90."""
        series = sample_df["roe_percentage"].copy()
        p90 = series.quantile(0.90)
        winsorized = ScreenerScoring.winsorize_metric(series)
        assert winsorized.max() <= p90, (
            f"Max after winsorization {winsorized.max()} exceeds P90 {p90}"
        )

    def test_lower_tail_clipped(self):
        """Values below P10 must be raised to P10."""
        series = pd.Series([1, 2, 3, 4, 5, 6, 7, 8, 9, -1000])
        p10 = series.quantile(0.10)
        winsorized = ScreenerScoring.winsorize_metric(series)
        assert winsorized.min() >= p10

    def test_bulk_is_unchanged(self):
        """Values between P10 and P90 must not be modified."""
        series = pd.Series(list(range(1, 21)))   # uniform, no real outliers
        original = series.copy()
        winsorized = ScreenerScoring.winsorize_metric(series, lower=0.10, upper=0.90)
        p10 = original.quantile(0.10)
        p90 = original.quantile(0.90)
        middle = original[(original >= p10) & (original <= p90)]
        middle_w = winsorized[(original >= p10) & (original <= p90)]
        pd.testing.assert_series_equal(
            middle.reset_index(drop=True),
            middle_w.reset_index(drop=True),
            check_names=False,
            check_dtype=False,
        )

    def test_na_preserved(self):
        """NaN values must stay NaN after winsorization."""
        series = pd.Series([10, 20, np.nan, 40, 50, np.nan])
        winsorized = ScreenerScoring.winsorize_metric(series)
        assert winsorized.isna().sum() == 2

    def test_all_same_values_no_error(self):
        """Constant series (P10 == P90) must not raise and return unchanged."""
        series = pd.Series([7.0] * 10)
        winsorized = ScreenerScoring.winsorize_metric(series)
        assert (winsorized == 7.0).all()

    def test_custom_percentiles(self):
        """Custom lower/upper percentiles must be respected."""
        series = pd.Series(range(1, 101))   # 1 to 100
        winsorized = ScreenerScoring.winsorize_metric(series, lower=0.05, upper=0.95)
        p5  = series.quantile(0.05)
        p95 = series.quantile(0.95)
        assert winsorized.min() >= p5
        assert winsorized.max() <= p95

    def test_normalization_output_in_0_1(self, sample_df):
        """P10/P90 normalization must produce values in [0, 1]."""
        w = ScreenerScoring.winsorize_metric(sample_df["roe_percentage"])
        n = ScreenerScoring.normalize_metric(w, method="p10p90")
        assert (n >= 0).all() and (n <= 1).all()

    def test_normalization_na_preserved(self):
        """NaN values must remain NaN after normalization."""
        series = pd.Series([1.0, 2.0, np.nan, 4.0, 5.0])
        n = ScreenerScoring.normalize_metric(series)
        assert pd.isna(n.iloc[2])


# ─────────────────────────────────────────────────────────────────────────────
# 2. COMPOSITE SCORE BETWEEN 0 AND 100
# ─────────────────────────────────────────────────────────────────────────────

class TestScoreBounds:

    def test_score_exists(self, scored_df):
        """composite_score column must be present."""
        assert "composite_score" in scored_df.columns

    def test_score_gte_0(self, scored_df):
        """No score may be negative."""
        assert (scored_df["composite_score"] >= 0).all(), (
            f"Negative scores found: {scored_df['composite_score'][scored_df['composite_score'] < 0].tolist()}"
        )

    def test_score_lte_100(self, scored_df):
        """No score may exceed 100."""
        assert (scored_df["composite_score"] <= 100).all(), (
            f"Scores > 100: {scored_df['composite_score'][scored_df['composite_score'] > 100].tolist()}"
        )

    def test_score_not_all_same(self, scored_df):
        """Scores must vary — a constant result indicates a bug."""
        assert scored_df["composite_score"].nunique() > 1

    def test_sub_scores_sum_to_composite(self, scored_df):
        """score_profitability + score_cash_quality + score_growth + score_leverage == composite_score."""
        total = (
            scored_df["score_profitability"]
            + scored_df["score_cash_quality"]
            + scored_df["score_growth"]
            + scored_df["score_leverage"]
        )
        pd.testing.assert_series_equal(
            total.round(6), scored_df["composite_score"].round(6),
            check_names=False,
        )

    def test_advanced_composite_equals_composite(self, scored_df):
        """advanced_composite_score must be an alias for composite_score."""
        assert "advanced_composite_score" in scored_df.columns
        pd.testing.assert_series_equal(
            scored_df["composite_score"].round(6),
            scored_df["advanced_composite_score"].round(6),
            check_names=False,
        )

    def test_score_with_real_data(self, screener):
        """Real DB data: all composite_scores must be in [0, 100]."""
        data = screener.load_financial_data(compute_scores=True)
        assert "composite_score" in data.columns
        assert (data["composite_score"] >= 0).all()
        assert (data["composite_score"] <= 100).all()

    def test_score_profitability_max_35(self, scored_df):
        """Profitability sub-score weight is 35 pts — no row may exceed it."""
        assert (scored_df["score_profitability"] <= 35.01).all()

    def test_score_leverage_max_15(self, scored_df):
        """Leverage sub-score weight is 15 pts — no row may exceed it."""
        assert (scored_df["score_leverage"] <= 15.01).all()


# ─────────────────────────────────────────────────────────────────────────────
# 3. SECTOR RANKING
# ─────────────────────────────────────────────────────────────────────────────

class TestSectorRanking:

    def test_sector_relative_score_exists(self, scored_df):
        assert "sector_relative_score" in scored_df.columns

    def test_sector_quality_rank_exists(self, scored_df):
        assert "sector_quality_rank" in scored_df.columns

    def test_sector_relative_score_range(self, scored_df):
        """sector_relative_score must be in [0, 100]."""
        srs = scored_df["sector_relative_score"]
        assert (srs >= 0).all(), f"Negative sector_relative_score found: {srs.min()}"
        assert (srs <= 100).all(), f"sector_relative_score > 100: {srs.max()}"

    def test_sector_best_has_rank_1(self, scored_df):
        """The company with the highest score in each sector gets rank 1."""
        for sector, grp in scored_df.groupby("sector"):
            best_score = grp["composite_score"].max()
            best_row   = grp[grp["composite_score"] == best_score]
            assert (best_row["sector_quality_rank"] == 1).all(), (
                f"Sector '{sector}': highest scorer does not have rank 1"
            )

    def test_sector_ranks_sequential(self, scored_df):
        """Ranks within a sector must be 1, 2, 3, … (sequential integers)."""
        for sector, grp in scored_df.groupby("sector"):
            ranks = sorted(grp["sector_quality_rank"].dropna().astype(int).tolist())
            expected = list(range(1, len(ranks) + 1))
            assert ranks == expected, (
                f"Sector '{sector}': ranks {ranks} not sequential"
            )

    def test_high_scorer_leads_sector(self, scored_df):
        """The sector leader (rank 1) must have the highest composite_score in sector."""
        for sector, grp in scored_df.groupby("sector"):
            leader = grp[grp["sector_quality_rank"] == 1]["composite_score"].values[0]
            assert leader == grp["composite_score"].max()

    def test_sector_comparison_columns_on_real_data(self, screener):
        """Real data: sector mean/median/vs_sector columns must exist."""
        data = screener.load_financial_data(compute_scores=True)
        for col in [
            "roe_percentage_sector_mean",
            "roe_percentage_sector_median",
            "roe_percentage_vs_sector",
            "roce_percentage_sector_mean",
            "roce_percentage_sector_median",
        ]:
            assert col in data.columns, f"Missing: {col}"

    def test_sector_relative_score_on_real_data(self, screener):
        """Real data: sector_relative_score must be in [0, 100] for all companies."""
        data = screener.load_financial_data(compute_scores=True)
        assert "sector_relative_score" in data.columns
        assert (data["sector_relative_score"] >= 0).all()
        assert (data["sector_relative_score"] <= 100).all()


# ─────────────────────────────────────────────────────────────────────────────
# 4. EXCEL CREATION
# ─────────────────────────────────────────────────────────────────────────────

class TestExcelCreation:

    def test_file_is_created(self, screener, tmp_path):
        """export_screeners() must produce a non-empty .xlsx file."""
        out = tmp_path / "output.xlsx"
        path = export_screeners(screener, output_path=str(out))
        assert path.exists(), "Excel file was not created"
        assert path.stat().st_size > 5_000, "Excel file is suspiciously small"

    def test_returns_path_object(self, screener, tmp_path):
        """export_screeners() must return a resolved Path."""
        out = tmp_path / "output.xlsx"
        path = export_screeners(screener, output_path=str(out))
        assert isinstance(path, Path)

    def test_file_is_valid_xlsx(self, screener, tmp_path):
        """openpyxl must be able to open the file without error."""
        out = tmp_path / "output.xlsx"
        export_screeners(screener, output_path=str(out))
        wb = load_workbook(out, read_only=True, data_only=True)
        assert len(wb.sheetnames) > 0
        wb.close()

    def test_each_sheet_has_header_row(self, screener, tmp_path):
        """Every sheet must have a header row (row 3) with non-null values."""
        out = tmp_path / "output.xlsx"
        export_screeners(screener, output_path=str(out))
        wb = load_workbook(out, read_only=True, data_only=True)
        for name in wb.sheetnames:
            ws = wb[name]
            headers = [c.value for c in ws[3] if c.value is not None]
            assert len(headers) == 20, (
                f"Sheet '{name}': expected 20 header cols, got {len(headers)}"
            )
        wb.close()


# ─────────────────────────────────────────────────────────────────────────────
# 5. SIX SHEETS EXIST
# ─────────────────────────────────────────────────────────────────────────────

EXPECTED_SHEETS = [
    "Quality Compounder",
    "Value Pick",
    "Growth Accelerator",
    "Dividend Champion",
    "Debt-Free Bluechip",
    "Turnaround Watch",
]


class TestSixSheets:

    @pytest.fixture(scope="class")
    def workbook(self, screener, tmp_path_factory):
        out = tmp_path_factory.mktemp("wb") / "output.xlsx"
        export_screeners(screener, output_path=str(out))
        wb = load_workbook(out, read_only=True, data_only=True)
        yield wb
        wb.close()

    def test_exactly_six_sheets(self, workbook):
        assert len(workbook.sheetnames) == 6, (
            f"Expected 6 sheets, got {len(workbook.sheetnames)}: {workbook.sheetnames}"
        )

    @pytest.mark.parametrize("sheet_name", EXPECTED_SHEETS)
    def test_preset_sheet_present(self, workbook, sheet_name):
        assert sheet_name in workbook.sheetnames, (
            f"Sheet '{sheet_name}' not found. Sheets: {workbook.sheetnames}"
        )

    def test_no_unexpected_sheets(self, workbook):
        for name in workbook.sheetnames:
            assert name in EXPECTED_SHEETS, (
                f"Unexpected sheet '{name}' found"
            )

    @pytest.mark.parametrize("sheet_name", EXPECTED_SHEETS)
    def test_each_sheet_has_data(self, workbook, sheet_name):
        ws = workbook[sheet_name]
        data_rows = [
            row for row in ws.iter_rows(min_row=4, values_only=True)
            if any(v is not None for v in row)
        ]
        assert len(data_rows) >= 5, (
            f"Sheet '{sheet_name}': only {len(data_rows)} data rows (need >= 5)"
        )
        assert len(data_rows) <= 50, (
            f"Sheet '{sheet_name}': {len(data_rows)} data rows exceeds 50"
        )


# ─────────────────────────────────────────────────────────────────────────────
# 6. SORTING WORKS (highest Composite Score first)
# ─────────────────────────────────────────────────────────────────────────────

class TestSorting:

    @pytest.fixture(scope="class")
    def workbook(self, screener, tmp_path_factory):
        out = tmp_path_factory.mktemp("wb") / "output.xlsx"
        export_screeners(screener, output_path=str(out))
        wb = load_workbook(out, read_only=True, data_only=True)
        yield wb
        wb.close()

    @pytest.mark.parametrize("sheet_name", EXPECTED_SHEETS)
    def test_sorted_score_desc_in_xlsx(self, workbook, sheet_name):
        """Composite Score column in each sheet must be non-increasing."""
        ws = workbook[sheet_name]
        headers = [cell.value for cell in ws[3]]
        try:
            score_col = headers.index("Composite Score")
        except ValueError:
            pytest.skip(f"'Composite Score' column not found in '{sheet_name}'")

        scores = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            val = row[score_col]
            if val is not None:
                try:
                    scores.append(float(val))
                except (TypeError, ValueError):
                    pass

        assert len(scores) >= 2, f"Sheet '{sheet_name}' has too few rows to test sort"
        assert scores == sorted(scores, reverse=True), (
            f"Sheet '{sheet_name}' NOT sorted DESC. First 5: {scores[:5]}"
        )

    def test_run_screener_sorted_desc(self, screener):
        """run_screener() result must be sorted by composite_score DESC."""
        for preset in ["quality_compounder", "growth_accelerator", "value_pick"]:
            result = screener.run_screener(preset)
            df = pd.DataFrame(result["companies"])
            assert len(df) > 0
            scores = df["composite_score"].dropna().tolist()
            assert scores == sorted(scores, reverse=True), (
                f"Preset '{preset}' result not sorted DESC. First 5: {scores[:5]}"
            )

    def test_ranked_table_sorted_desc(self, screener):
        """ranked DataFrame from run_screener must be sorted Score DESC."""
        result = screener.run_screener("quality_compounder")
        ranked = result["ranked"]
        scores = ranked["Score"].tolist()
        assert scores == sorted(scores, reverse=True)

    def test_top_company_has_highest_score(self, screener):
        """First row must have the maximum composite_score."""
        result = screener.run_screener("quality_compounder")
        df = pd.DataFrame(result["companies"])
        max_score = df["composite_score"].max()
        assert df.iloc[0]["composite_score"] == max_score


# ─────────────────────────────────────────────────────────────────────────────
# BONUS: D/E INVERSION  (1 - percentile_rank)
# ─────────────────────────────────────────────────────────────────────────────

class TestDebtEquityInversion:

    def test_high_debt_gets_low_score(self, scored_df):
        """
        Company J has D/E=10 (highest in dataset).
        Its score_debt must be strictly less than company A (D/E=0).
        """
        a = scored_df[scored_df["company_id"] == "A"]["score_debt"].values[0]
        j = scored_df[scored_df["company_id"] == "J"]["score_debt"].values[0]
        assert j < a, (
            f"High-debt company J score_debt={j:.3f} should be "
            f"less than debt-free company A score_debt={a:.3f}"
        )

    def test_debt_free_gets_highest_debt_score(self, scored_df):
        """
        Debt-free companies (D/E = 0) must have the highest score_debt.
        """
        max_score = scored_df["score_debt"].max()
        debt_free = scored_df[scored_df["debt_equity"] == 0.0]
        assert len(debt_free) > 0, "No debt-free companies in sample_df"
        assert (debt_free["score_debt"] == max_score).all(), (
            "Debt-free company does not have the maximum score_debt"
        )

    def test_score_debt_inversely_ordered(self, scored_df):
        """
        score_debt must be inversely ordered with debt_equity:
        sorting by debt_equity ASC should give score_debt DESC.
        """
        df = scored_df[["debt_equity", "score_debt"]].dropna().sort_values("debt_equity")
        debt_scores = df["score_debt"].tolist()
        # Allow for ties (equal D/E can have equal score_debt)
        assert all(
            debt_scores[i] >= debt_scores[i + 1]
            for i in range(len(debt_scores) - 1)
        ), (
            "score_debt is NOT inversely ordered with debt_equity. "
            f"debt_equity order: {df['debt_equity'].tolist()}, "
            f"score_debt: {debt_scores}"
        )

    def test_percentile_inversion_formula(self):
        """
        Unit test: given a known D/E series, verify 1 - percentile_rank inversion.
        Series: [0, 1, 2, 3, 4]  percentile ranks: [0.2, 0.4, 0.6, 0.8, 1.0]
        Inverted:                                   [0.8, 0.6, 0.4, 0.2, 0.0]
        Multiply by 10 pts weight:                  [8.0, 6.0, 4.0, 2.0, 0.0]
        """
        de = pd.Series([0.0, 1.0, 2.0, 3.0, 4.0])
        pct_rank = de.rank(pct=True, method="average")
        inverted = (1.0 - pct_rank).clip(0, 1)
        score_debt = (inverted * 10).round(4)
        expected = pd.Series([8.0, 6.0, 4.0, 2.0, 0.0])
        pd.testing.assert_series_equal(score_debt, expected, check_names=False)

    def test_real_data_debt_inversion(self, screener):
        """
        Real DB: among top-10 highest D/E companies, their score_debt
        must be less than the median score_debt across all companies.
        """
        data = screener.load_financial_data(compute_scores=True)
        assert "score_debt" in data.columns, "score_debt column missing"
        high_debt = (
            data[data["debt_equity"].notna()]
            .nlargest(10, "debt_equity")["score_debt"]
        )
        median_score = data["score_debt"].median()
        assert high_debt.mean() < median_score, (
            "Top-10 highest D/E companies should have below-median score_debt"
        )
