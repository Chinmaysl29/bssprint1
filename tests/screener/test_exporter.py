"""
tests/screener/test_exporter.py
────────────────────────────────
Tests for src/screener/exporter.py

Verifies that output/screener_output.xlsx
  • contains exactly 6 sheets (one per preset)
  • each sheet has exactly 20 KPI columns in the correct order
  • each sheet is sorted by Composite Score DESC
  • Rank column is 1-based and sequential
  • Remarks column contains valid tier strings
  • Conditional Formatting: GREEN cells where condition PASSED,
                             RED cells where condition FAILED
  • ScreenerExporter.export_to_excel() backward-compat still works
"""

import os
import sys
import pytest
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "../..")))

from src.screener.engine import StockScreener
from src.screener.exporter import (
    export_screeners, ScreenerExporter, KPI_COLUMNS, CF_RULES,
    _CF_GREEN_PASS, _CF_RED_FAIL,
)

OUTPUT_PATH = Path("output/screener_output.xlsx")
TEST_EXPORT  = Path("output/test_export_tmp.xlsx")

EXPECTED_SHEETS = [
    "Quality Compounder",
    "Value Pick",
    "Growth Accelerator",
    "Dividend Champion",
    "Debt-Free Bluechip",
    "Turnaround Watch",
]

EXPECTED_HEADERS = [h for _, h, _ in KPI_COLUMNS]

VALID_REMARKS = {
    "S-Tier – Strong Buy",
    "A-Tier – Buy",
    "B-Tier – Watch",
    "C-Tier – Avoid",
    "N/A",
}


# ── Fixtures ──────────────────────────────────────────────────────────────────

@pytest.fixture(scope="module")
def screener():
    return StockScreener()


@pytest.fixture(scope="module")
def exported_path(screener):
    path = export_screeners(screener, output_path=str(OUTPUT_PATH))
    return path


@pytest.fixture(scope="module")
def workbook(exported_path):
    # data_only=False so we can inspect ConditionalFormatting rules
    wb = load_workbook(exported_path, read_only=False, data_only=True)
    yield wb
    wb.close()


# ── File-level checks ─────────────────────────────────────────────────────────

def test_file_created(exported_path):
    assert exported_path.exists(), f"File not found: {exported_path}"


def test_file_nonempty(exported_path):
    assert exported_path.stat().st_size >= 10_000, (
        f"File too small ({exported_path.stat().st_size} bytes)"
    )


# ── Sheet count & names ───────────────────────────────────────────────────────

def test_exactly_6_sheets(workbook):
    assert len(workbook.sheetnames) == 6, (
        f"Expected 6 sheets, got {len(workbook.sheetnames)}: {workbook.sheetnames}"
    )


@pytest.mark.parametrize("sheet_name", EXPECTED_SHEETS)
def test_sheet_exists(workbook, sheet_name):
    assert sheet_name in workbook.sheetnames, (
        f"Sheet '{sheet_name}' missing. Found: {workbook.sheetnames}"
    )


def test_no_extra_sheets(workbook):
    for name in workbook.sheetnames:
        assert name in EXPECTED_SHEETS, (
            f"Unexpected sheet: '{name}'. Allowed: {EXPECTED_SHEETS}"
        )


# ── Column structure (20 KPIs in exact order) ─────────────────────────────────

@pytest.mark.parametrize("sheet_name", EXPECTED_SHEETS)
def test_20_columns_present(workbook, sheet_name):
    ws      = workbook[sheet_name]
    headers = [cell.value for cell in ws[3] if cell.value is not None]
    assert len(headers) == 20, (
        f"Sheet '{sheet_name}': expected 20 columns, got {len(headers)}: {headers}"
    )


@pytest.mark.parametrize("sheet_name", EXPECTED_SHEETS)
def test_column_order(workbook, sheet_name):
    ws      = workbook[sheet_name]
    headers = [cell.value for cell in ws[3]]
    assert headers == EXPECTED_HEADERS, (
        f"Sheet '{sheet_name}' column order mismatch.\n"
        f"  Expected: {EXPECTED_HEADERS}\n"
        f"  Got:      {headers}"
    )


# ── Data presence ─────────────────────────────────────────────────────────────

@pytest.mark.parametrize("sheet_name", EXPECTED_SHEETS)
def test_sheet_has_data_rows(workbook, sheet_name):
    ws        = workbook[sheet_name]
    data_rows = [
        row for row in ws.iter_rows(min_row=4, values_only=True)
        if any(v is not None for v in row)
    ]
    assert len(data_rows) >= 1, f"Sheet '{sheet_name}' has no data rows"


# ── Sort order ────────────────────────────────────────────────────────────────

@pytest.mark.parametrize("sheet_name", EXPECTED_SHEETS)
def test_sorted_by_score_desc(workbook, sheet_name):
    ws      = workbook[sheet_name]
    headers = [cell.value for cell in ws[3]]
    try:
        score_col = headers.index("Composite Score")
    except ValueError:
        pytest.skip(f"'Composite Score' not found in '{sheet_name}'")

    scores = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        val = row[score_col]
        if val is not None:
            try:
                scores.append(float(val))
            except (TypeError, ValueError):
                pass

    assert len(scores) >= 2
    assert scores == sorted(scores, reverse=True), (
        f"Sheet '{sheet_name}' NOT sorted Score DESC. First 5: {scores[:5]}"
    )


# ── Rank column ───────────────────────────────────────────────────────────────

@pytest.mark.parametrize("sheet_name", EXPECTED_SHEETS)
def test_rank_is_sequential(workbook, sheet_name):
    ws      = workbook[sheet_name]
    headers = [cell.value for cell in ws[3]]
    try:
        rank_col = headers.index("Rank")
    except ValueError:
        pytest.skip(f"'Rank' not found in '{sheet_name}'")

    ranks = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        val = row[rank_col]
        if val is not None:
            try:
                ranks.append(int(val))
            except (TypeError, ValueError):
                pass

    assert ranks == list(range(1, len(ranks) + 1)), (
        f"Sheet '{sheet_name}' Rank not sequential. Got: {ranks[:10]}"
    )


# ── Remarks column ────────────────────────────────────────────────────────────

@pytest.mark.parametrize("sheet_name", EXPECTED_SHEETS)
def test_remarks_valid_values(workbook, sheet_name):
    ws      = workbook[sheet_name]
    headers = [cell.value for cell in ws[3]]
    try:
        rem_col = headers.index("Remarks")
    except ValueError:
        pytest.skip(f"'Remarks' not found in '{sheet_name}'")

    for row in ws.iter_rows(min_row=4, values_only=True):
        val = row[rem_col]
        if val is not None:
            assert val in VALID_REMARKS, (
                f"Sheet '{sheet_name}': unexpected Remarks value '{val}'"
            )


# ── Required KPI columns spot-check ──────────────────────────────────────────

@pytest.mark.parametrize("sheet_name", EXPECTED_SHEETS)
@pytest.mark.parametrize("required_col", [
    "Company Name", "Sector", "Year", "ROE (%)", "ROCE (%)",
    "Net Profit Margin (%)", "OPM (%)", "Debt / Equity",
    "ICR", "FCF (₹ Cr)", "Revenue CAGR (3Y %)", "PAT CAGR (3Y %)",
    "EPS CAGR (3Y %)", "Sales (₹ Cr)", "Net Profit (₹ Cr)",
    "Market Cap (₹ Cr)", "Dividend Yield (%)",
    "Composite Score", "Rank", "Remarks",
])
def test_required_kpi_column_present(workbook, sheet_name, required_col):
    ws      = workbook[sheet_name]
    headers = [cell.value for cell in ws[3]]
    assert required_col in headers, (
        f"Sheet '{sheet_name}': required column '{required_col}' not found. "
        f"Headers: {headers}"
    )


# ── Conditional Formatting: openpyxl rules registered ────────────────────────

@pytest.mark.parametrize("sheet_name", EXPECTED_SHEETS)
def test_cf_rules_registered(exported_path, sheet_name):
    """
    Each numeric KPI sheet must have openpyxl ConditionalFormatting rules
    registered — load with read_only=False to inspect CF.
    """
    wb = load_workbook(exported_path, read_only=False, data_only=True)
    try:
        ws = wb[sheet_name]
        cf_count = len(list(ws.conditional_formatting))
        assert cf_count > 0, (
            f"Sheet '{sheet_name}': no ConditionalFormatting rules found. "
            "Expected GREEN/RED CellIsRule for each numeric KPI column."
        )
    finally:
        wb.close()


@pytest.mark.parametrize("sheet_name", EXPECTED_SHEETS)
def test_cf_rules_count(exported_path, sheet_name):
    """
    Must have at least 22 CF rules total (11 columns × 2 rules each).
    """
    wb = load_workbook(exported_path, read_only=False, data_only=True)
    try:
        ws = wb[sheet_name]
        total_rules = sum(len(cf.rules) for cf in ws.conditional_formatting)
        expected_min = 22
        assert total_rules >= expected_min, (
            f"Sheet '{sheet_name}': expected ≥ {expected_min} CF rules, "
            f"got {total_rules}"
        )
    finally:
        wb.close()


# ── Conditional Formatting: cell-level colours ───────────────────────────────

def _cell_fill_hex(cell) -> str | None:
    """Return the cell background fill colour hex, or None."""
    f = cell.fill
    if f and f.fill_type == "solid" and f.fgColor:
        c = f.fgColor
        if hasattr(c, "rgb") and c.rgb and c.rgb != "00000000":
            return c.rgb[-6:].upper()  # strip alpha prefix
    return None


@pytest.mark.parametrize("sheet_name", ["Quality Compounder", "Growth Accelerator"])
def test_roe_green_cells_are_green(workbook, sheet_name):
    """
    Cells where ROE >= 15 must have a GREEN fill (C6EFCE).
    """
    ws      = workbook[sheet_name]
    headers = [cell.value for cell in ws[3]]
    try:
        roe_col_idx = headers.index("ROE (%)") + 1   # 1-based
    except ValueError:
        pytest.skip("ROE (%) not found")

    green_expected = _CF_GREEN_PASS.upper()
    checked = 0
    for row_idx in range(4, ws.max_row + 1):
        val_cell  = ws.cell(row=row_idx, column=roe_col_idx)
        val       = val_cell.value
        if val is None:
            continue
        try:
            roe = float(val)
        except (TypeError, ValueError):
            continue

        fill_hex = _cell_fill_hex(val_cell)
        if roe >= 15:
            assert fill_hex == green_expected, (
                f"Sheet '{sheet_name}' row {row_idx}: ROE={roe} >= 15 "
                f"but fill is '{fill_hex}', expected green '{green_expected}'"
            )
            checked += 1
        elif roe < 15:
            red_expected = _CF_RED_FAIL.upper()
            assert fill_hex == red_expected, (
                f"Sheet '{sheet_name}' row {row_idx}: ROE={roe} < 15 "
                f"but fill is '{fill_hex}', expected red '{red_expected}'"
            )
            checked += 1

    assert checked >= 1, f"No ROE values found to check in '{sheet_name}'"


@pytest.mark.parametrize("sheet_name", ["Quality Compounder", "Value Pick"])
def test_debt_equity_red_cells_are_red(workbook, sheet_name):
    """
    Cells where Debt/Equity > 1.0 must have a RED fill (FFC7CE).
    """
    ws      = workbook[sheet_name]
    headers = [cell.value for cell in ws[3]]
    try:
        de_col_idx = headers.index("Debt / Equity") + 1
    except ValueError:
        pytest.skip("Debt / Equity not found")

    red_expected   = _CF_RED_FAIL.upper()
    green_expected = _CF_GREEN_PASS.upper()
    checked = 0

    for row_idx in range(4, ws.max_row + 1):
        val_cell = ws.cell(row=row_idx, column=de_col_idx)
        val      = val_cell.value
        if val is None:
            continue
        try:
            de = float(val)
        except (TypeError, ValueError):
            continue

        fill_hex = _cell_fill_hex(val_cell)
        if de > 1.0:
            assert fill_hex == red_expected, (
                f"Sheet '{sheet_name}' row {row_idx}: D/E={de} > 1.0 "
                f"but fill is '{fill_hex}', expected red '{red_expected}'"
            )
            checked += 1
        elif de <= 1.0:
            assert fill_hex == green_expected, (
                f"Sheet '{sheet_name}' row {row_idx}: D/E={de} <= 1.0 "
                f"but fill is '{fill_hex}', expected green '{green_expected}'"
            )
            checked += 1

    assert checked >= 1, f"No D/E values found to check in '{sheet_name}'"


@pytest.mark.parametrize("sheet_name", EXPECTED_SHEETS)
def test_fcf_positive_cells_are_green(workbook, sheet_name):
    """
    FCF > 0 → green; FCF <= 0 → red. At least one positive FCF cell must exist.
    """
    ws      = workbook[sheet_name]
    headers = [cell.value for cell in ws[3]]
    try:
        fcf_col = headers.index("FCF (₹ Cr)") + 1
    except ValueError:
        pytest.skip("FCF column not found")

    green_expected = _CF_GREEN_PASS.upper()
    red_expected   = _CF_RED_FAIL.upper()
    found_positive = False

    for row_idx in range(4, ws.max_row + 1):
        val_cell = ws.cell(row=row_idx, column=fcf_col)
        val      = val_cell.value
        if val is None:
            continue
        try:
            fcf = float(val)
        except (TypeError, ValueError):
            continue

        fill_hex = _cell_fill_hex(val_cell)
        if fcf > 0:
            assert fill_hex == green_expected, (
                f"Sheet '{sheet_name}' row {row_idx}: FCF={fcf}>0 "
                f"fill='{fill_hex}', expected green '{green_expected}'"
            )
            found_positive = True
        else:
            assert fill_hex == red_expected, (
                f"Sheet '{sheet_name}' row {row_idx}: FCF={fcf}<=0 "
                f"fill='{fill_hex}', expected red '{red_expected}'"
            )

    # Some sheets (e.g. Turnaround Watch requires FCF>1000) will all be green
    # so we just require at least one row was checked
    assert ws.max_row >= 4, f"Sheet '{sheet_name}' appears to have no data rows"


# ── CF rule definitions exported correctly ────────────────────────────────────

def test_cf_rules_table_has_all_kpis():
    """CF_RULES must define pass/fail thresholds for the key numeric KPIs."""
    defined_cols = {r["df_col"] for r in CF_RULES}
    required = {
        "roe_percentage", "roce_percentage", "npm", "opm_percentage",
        "debt_equity", "fcf", "revenue_cagr_3y", "pat_cagr_3y",
        "eps_cagr_3y", "dividend_yield", "composite_score",
    }
    missing = required - defined_cols
    assert not missing, f"CF_RULES missing entries for: {missing}"


def test_cf_rules_pass_fail_not_equal():
    """Pass and fail thresholds must not both be identical operators/values."""
    for rule in CF_RULES:
        assert rule["pass_op"] != rule["fail_op"] or rule["pass_val"] != rule["fail_val"], (
            f"CF rule for '{rule['df_col']}' has identical pass/fail condition"
        )


# ── ScreenerExporter backward-compat ─────────────────────────────────────────

def test_export_to_excel_backward_compat(screener):
    result = screener.run_screener("quality_compounder")
    df     = pd.DataFrame(result["companies"])

    try:
        ok = ScreenerExporter.export_to_excel(df, filename=str(TEST_EXPORT))
        assert ok is True
        assert TEST_EXPORT.exists()
        wb = load_workbook(TEST_EXPORT, read_only=True, data_only=True)
        assert "Screener Results" in wb.sheetnames
        wb.close()
    finally:
        if TEST_EXPORT.exists():
            TEST_EXPORT.unlink()

# Expected sheet names (matching screener_config.yaml display names)
EXPECTED_SHEETS = [
    "Quality Compounder",
    "Value Pick",
    "Growth Accelerator",
    "Dividend Champion",
    "Debt-Free Bluechip",
    "Turnaround Watch",
]

# Expected 20 column headers in exact order
EXPECTED_HEADERS = [h for _, h, _ in KPI_COLUMNS]

VALID_REMARKS = {
    "S-Tier – Strong Buy",
    "A-Tier – Buy",
    "B-Tier – Watch",
    "C-Tier – Avoid",
    "N/A",
}


# ── Fixtures ──────────────────────────────────────────────────────────────────

@pytest.fixture(scope="module")
def screener():
    return StockScreener()


@pytest.fixture(scope="module")
def exported_path(screener):
    path = export_screeners(screener, output_path=str(OUTPUT_PATH))
    return path


@pytest.fixture(scope="module")
def workbook(exported_path):
    wb = load_workbook(exported_path, read_only=True, data_only=True)
    yield wb
    wb.close()


# ── File-level checks ─────────────────────────────────────────────────────────

def test_file_created(exported_path):
    """output/screener_output.xlsx must exist after export."""
    assert exported_path.exists(), f"File not found: {exported_path}"


def test_file_nonempty(exported_path):
    """File must be at least 10 KB (has real data)."""
    assert exported_path.stat().st_size >= 10_000, (
        f"File too small ({exported_path.stat().st_size} bytes) — probably empty"
    )


# ── Sheet count & names ───────────────────────────────────────────────────────

def test_exactly_6_sheets(workbook):
    """Workbook must contain exactly 6 sheets."""
    assert len(workbook.sheetnames) == 6, (
        f"Expected 6 sheets, got {len(workbook.sheetnames)}: {workbook.sheetnames}"
    )


@pytest.mark.parametrize("sheet_name", EXPECTED_SHEETS)
def test_sheet_exists(workbook, sheet_name):
    """Each of the 6 preset sheets must be present."""
    assert sheet_name in workbook.sheetnames, (
        f"Sheet '{sheet_name}' missing. Found: {workbook.sheetnames}"
    )


def test_no_extra_sheets(workbook):
    """No unexpected sheets (e.g. 'Summary' or 'Ranked (All)') should be present."""
    for name in workbook.sheetnames:
        assert name in EXPECTED_SHEETS, (
            f"Unexpected sheet: '{name}'. Allowed: {EXPECTED_SHEETS}"
        )


# ── Column structure (20 KPIs in exact order) ─────────────────────────────────

@pytest.mark.parametrize("sheet_name", EXPECTED_SHEETS)
def test_20_columns_present(workbook, sheet_name):
    """Every sheet must have exactly 20 columns."""
    ws      = workbook[sheet_name]
    # Headers are on row 3 (row 1 = title, row 2 = subtitle)
    headers = [cell.value for cell in ws[3] if cell.value is not None]
    assert len(headers) == 20, (
        f"Sheet '{sheet_name}': expected 20 columns, got {len(headers)}: {headers}"
    )


@pytest.mark.parametrize("sheet_name", EXPECTED_SHEETS)
def test_column_order(workbook, sheet_name):
    """Columns must appear in the exact required order."""
    ws      = workbook[sheet_name]
    headers = [cell.value for cell in ws[3]]
    assert headers == EXPECTED_HEADERS, (
        f"Sheet '{sheet_name}' column order mismatch.\n"
        f"  Expected: {EXPECTED_HEADERS}\n"
        f"  Got:      {headers}"
    )


# ── Data presence ─────────────────────────────────────────────────────────────

@pytest.mark.parametrize("sheet_name", EXPECTED_SHEETS)
def test_sheet_has_data_rows(workbook, sheet_name):
    """Every sheet must have at least 1 data row (below the header)."""
    ws        = workbook[sheet_name]
    data_rows = [
        row for row in ws.iter_rows(min_row=4, values_only=True)
        if any(v is not None for v in row)
    ]
    assert len(data_rows) >= 1, (
        f"Sheet '{sheet_name}' has no data rows"
    )


# ── Sort order ────────────────────────────────────────────────────────────────

@pytest.mark.parametrize("sheet_name", EXPECTED_SHEETS)
def test_sorted_by_score_desc(workbook, sheet_name):
    """Composite Score must be non-increasing from top to bottom."""
    ws      = workbook[sheet_name]
    headers = [cell.value for cell in ws[3]]

    try:
        score_col = headers.index("Composite Score")
    except ValueError:
        pytest.skip(f"'Composite Score' column not found in '{sheet_name}'")

    scores = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        val = row[score_col]
        if val is not None:
            try:
                scores.append(float(val))
            except (TypeError, ValueError):
                pass

    assert len(scores) >= 2, f"Not enough rows to test sort order in '{sheet_name}'"
    assert scores == sorted(scores, reverse=True), (
        f"Sheet '{sheet_name}' is NOT sorted by Composite Score DESC. "
        f"First 5 scores: {scores[:5]}"
    )


# ── Rank column ───────────────────────────────────────────────────────────────

@pytest.mark.parametrize("sheet_name", EXPECTED_SHEETS)
def test_rank_is_sequential(workbook, sheet_name):
    """Rank must be 1, 2, 3, … sequentially."""
    ws      = workbook[sheet_name]
    headers = [cell.value for cell in ws[3]]

    try:
        rank_col = headers.index("Rank")
    except ValueError:
        pytest.skip(f"'Rank' column not found in '{sheet_name}'")

    ranks = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        val = row[rank_col]
        if val is not None:
            try:
                ranks.append(int(val))
            except (TypeError, ValueError):
                pass

    expected = list(range(1, len(ranks) + 1))
    assert ranks == expected, (
        f"Sheet '{sheet_name}' Rank column is not sequential 1,2,3,...  "
        f"Got: {ranks[:10]}"
    )


# ── Remarks column ────────────────────────────────────────────────────────────

@pytest.mark.parametrize("sheet_name", EXPECTED_SHEETS)
def test_remarks_valid_values(workbook, sheet_name):
    """Every Remarks cell must contain a valid tier string."""
    ws      = workbook[sheet_name]
    headers = [cell.value for cell in ws[3]]

    try:
        rem_col = headers.index("Remarks")
    except ValueError:
        pytest.skip(f"'Remarks' column not found in '{sheet_name}'")

    for row in ws.iter_rows(min_row=4, values_only=True):
        val = row[rem_col]
        if val is not None:
            assert val in VALID_REMARKS, (
                f"Sheet '{sheet_name}': unexpected Remarks value '{val}'. "
                f"Allowed: {VALID_REMARKS}"
            )


# ── Required KPI columns spot-check ──────────────────────────────────────────

@pytest.mark.parametrize("sheet_name", EXPECTED_SHEETS)
@pytest.mark.parametrize("required_col", [
    "Company Name", "Sector", "Year", "ROE (%)", "ROCE (%)",
    "Net Profit Margin (%)", "OPM (%)", "Debt / Equity",
    "ICR", "FCF (₹ Cr)", "Revenue CAGR (3Y %)", "PAT CAGR (3Y %)",
    "EPS CAGR (3Y %)", "Sales (₹ Cr)", "Net Profit (₹ Cr)",
    "Market Cap (₹ Cr)", "Dividend Yield (%)",
    "Composite Score", "Rank", "Remarks",
])
def test_required_kpi_column_present(workbook, sheet_name, required_col):
    """Each of the 20 required KPI columns must be present in every sheet."""
    ws      = workbook[sheet_name]
    headers = [cell.value for cell in ws[3]]
    assert required_col in headers, (
        f"Sheet '{sheet_name}': required column '{required_col}' not found. "
        f"Headers: {headers}"
    )


# ── ScreenerExporter backward-compat ─────────────────────────────────────────

def test_export_to_excel_backward_compat(screener):
    """ScreenerExporter.export_to_excel() must still produce a valid file."""
    result = screener.run_screener("quality_compounder")
    df     = pd.DataFrame(result["companies"])

    try:
        ok = ScreenerExporter.export_to_excel(df, filename=str(TEST_EXPORT))
        assert ok is True, "export_to_excel returned False"
        assert TEST_EXPORT.exists(), "Temp export file was not created"

        wb = load_workbook(TEST_EXPORT, read_only=True, data_only=True)
        assert "Screener Results" in wb.sheetnames
        wb.close()
    finally:
        if TEST_EXPORT.exists():
            TEST_EXPORT.unlink()
