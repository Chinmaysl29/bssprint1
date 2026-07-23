from pathlib import Path

from openpyxl import Workbook
from openpyxl import load_workbook

from src.analytics.peer_report import (
    PERCENTILE_METRICS,
    REPORT_COLUMNS,
    REPORT_SHEETS,
    add_peer_median,
    apply_formatting,
    export_excel,
    generate_peer_sheets,
    load_peer_data,
)


def test_load_peer_data_has_records():
    peer_df = load_peer_data()

    assert not peer_df.empty
    assert list(peer_df.columns) == REPORT_COLUMNS + ["report_sheet"]
    assert peer_df.loc[peer_df["company_id"].eq("TCS"), "Composite Score"].notna().any()
    percentile_columns = [column for column in REPORT_COLUMNS if column.endswith("_percentile")]
    assert len(percentile_columns) == 10
    assert set(percentile_columns) == set(PERCENTILE_METRICS.values())
    assert {"ROE_percentile", "ROCE_percentile", "Debt_percentile"}.issubset(peer_df.columns)


def test_add_peer_median_adds_expected_columns():
    peer_df = load_peer_data().head(200)
    report_df = add_peer_median(peer_df)

    assert list(report_df.columns) == list(peer_df.columns)


def test_generate_peer_sheets_returns_peer_group_tabs():
    sheets = generate_peer_sheets(load_peer_data())

    assert list(sheets) == REPORT_SHEETS
    assert len(sheets) == 11
    for sheet_df in sheets.values():
        assert list(sheet_df.columns) == REPORT_COLUMNS
        assert sheet_df.iloc[-1]["company_name"] == "Peer Median"


def test_export_excel_creates_exact_11_sheet_workbook(tmp_path):
    output_path = tmp_path / "peer_comparison.xlsx"

    exported_path = export_excel(output_path=output_path)

    assert exported_path == output_path
    assert Path(exported_path).exists()

    workbook = load_workbook(exported_path, read_only=False)
    try:
        assert workbook.sheetnames == REPORT_SHEETS
        assert len(workbook.sheetnames) == 11
        worksheet = workbook["IT Services"]
        headers = [cell.value for cell in worksheet[1]]
        assert headers == REPORT_COLUMNS
        assert len([header for header in headers if str(header).endswith("_percentile")]) == 10
        assert worksheet.max_row > 1
        assert worksheet.cell(row=worksheet.max_row, column=headers.index("company_name") + 1).value == "Peer Median"
        assert worksheet.cell(row=worksheet.max_row, column=headers.index("ROE") + 1).value is not None
        assert worksheet.cell(row=worksheet.max_row, column=headers.index("company_name") + 1).fill.fgColor.rgb == "00FCD34D"
        assert worksheet.freeze_panes == "A2"
    finally:
        workbook.close()


def test_percentile_color_formatting():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["ROE_percentile"])
    worksheet.append([80])
    worksheet.append([50])
    worksheet.append([25])

    apply_formatting(workbook)

    assert worksheet["A2"].fill.fgColor.rgb == "00C6EFCE"
    assert worksheet["A3"].fill.fgColor.rgb == "00FFEB9C"
    assert worksheet["A4"].fill.fgColor.rgb == "00FFC7CE"


def test_benchmark_row_color_formatting():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["company_name", "ROE_percentile", "Peer Rank"])
    worksheet.append(["Benchmark", 80, 1])
    worksheet.append(["Other", 50, 2])

    apply_formatting(workbook)

    assert worksheet["A2"].fill.fgColor.rgb == "00FFD966"
    assert worksheet["C2"].fill.fgColor.rgb == "00FFD966"
    assert worksheet["B2"].fill.fgColor.rgb == "00C6EFCE"
    assert worksheet["A3"].fill.fgColor.rgb != "00FFD966"
